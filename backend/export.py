"""
Excel export functionality.

This module provides functions to export category and entry data
to Excel workbooks with proper formatting.
"""

from typing import List, Optional
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlmodel import select

from .db import get_session
from .models import Category, Entry
from .utils import sanitize_excel_sheet_title
from .logger import get_logger


logger = get_logger("export")


def _format_date_to_german(date_str: str) -> str:
    """
    Convert date from YYYY-MM format to dd.MM.yyyy format.
    
    Args:
        date_str: Date string in YYYY-MM format
        
    Returns:
        Date string in dd.MM.yyyy format (with 01 as day)
    """
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m")
        return date_obj.strftime("01.%m.%Y")
    except ValueError:
        # If format is unexpected, return as-is
        return date_str


def _save_workbook(wb: Workbook) -> BytesIO:
    """
    Save workbook to BytesIO buffer.
    
    Args:
        wb: Workbook to save
        
    Returns:
        BytesIO buffer containing workbook data
    """
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _apply_header_formatting(ws, headers: List[str]) -> None:
    """
    Apply bold formatting to header row.
    
    Args:
        ws: Worksheet
        headers: List of header strings
    """
    bold_font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold_font


def _auto_adjust_column_widths(ws) -> None:
    """
    Automatically adjust column widths based on content.
    
    Args:
        ws: Worksheet to adjust
    """
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = max_length + 2
        col_letter = column_cells[0].column_letter
        ws.column_dimensions[col_letter].width = adjusted_width


def generate_workbook(
    category_ids: Optional[List[int]] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
) -> BytesIO:
    """
    Generate Excel workbook with category data.
    
    Creates one sheet per category with all entries and metadata.
    When exporting all categories or multiple categories, also creates
    a summary sheet with category metadata.
    
    Args:
        category_ids: List of category IDs to export (None = all)
        from_date: Start date filter (YYYY-MM)
        to_date: End date filter (YYYY-MM)
        
    Returns:
        BytesIO buffer containing Excel file
        
    Raises:
        Exception: If database operations fail
    """
    logger.info(
        f"Generating workbook: categories={category_ids}, "
        f"from={from_date}, to={to_date}"
    )
    
    wb = Workbook()
    
    try:
        with get_session() as s:
            # Fetch categories
            if category_ids:
                cats = s.exec(
                    select(Category).where(Category.id.in_(category_ids))
                ).all()
            else:
                cats = s.exec(select(Category)).all()
            
            logger.debug(f"Exporting {len(cats)} categories")
            
            # Handle case with no categories
            if not cats:
                ws = wb.active
                ws.title = "No categories"
                ws.append(["Info"])
                ws.append(["No categories found for export"])
                logger.warning("No categories found for export")
                return _save_workbook(wb)
            
            # Remove default sheet if we have categories
            if "Sheet" in wb.sheetnames and len(cats) > 0:
                wb.remove(wb["Sheet"])
            
            # Create summary sheet if exporting all categories or multiple categories
            # (but not for single category export)
            is_multi_category_export = category_ids is None or len(category_ids) > 1
            if is_multi_category_export:
                summary_ws = wb.create_sheet(title="Übersicht", index=0)
                summary_headers = ["Kategorie", "Typ", "Einheit", "Auto-Erstellung"]
                summary_ws.append(summary_headers)
                _apply_header_formatting(summary_ws, summary_headers)
                
                # Add category metadata
                for cat in cats:
                    type_display = "Sparkategorie" if cat.type == "sparen" else "Normal"
                    auto_create_display = "Ja" if cat.auto_create else "Nein"
                    summary_ws.append([
                        cat.name,
                        type_display,
                        cat.unit or "",
                        auto_create_display
                    ])
                
                _auto_adjust_column_widths(summary_ws)
                logger.debug("Summary sheet created with category metadata")
            
            # Create sheet for each category
            for cat in cats:
                sheet_title = sanitize_excel_sheet_title(
                    cat.name or f"cat_{cat.id}"
                )
                ws = wb.create_sheet(title=sheet_title)
                
                # Headers
                headers = [
                    "Kategorie", "Datum", "Wert", "Einzahlung",
                    "Einheit", "Kommentar"
                ]
                ws.append(headers)
                _apply_header_formatting(ws, headers)
                
                # Fetch entries with filters
                stmt = select(Entry).where(Entry.category_id == cat.id)
                if from_date:
                    stmt = stmt.where(Entry.date >= from_date)
                if to_date:
                    stmt = stmt.where(Entry.date <= to_date)
                stmt = stmt.order_by(Entry.date)
                
                entries = s.exec(stmt).all()
                logger.debug(f"Exporting {len(entries)} entries for category {cat.name}")
                
                # Add entry rows
                for entry in entries:
                    ws.append([
                        cat.name,
                        _format_date_to_german(entry.date),
                        entry.value,
                        entry.deposit if entry.deposit is not None else "",
                        cat.unit or "",
                        entry.comment or ""
                    ])
                
                # Auto-adjust column widths
                _auto_adjust_column_widths(ws)
            
            logger.info(f"Workbook generated successfully with {len(cats)} sheets")
            return _save_workbook(wb)
    
    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        raise